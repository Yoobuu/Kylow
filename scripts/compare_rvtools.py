import sys
import openpyxl
from pathlib import Path

# Sheets supported by rvtools_compat layer that should match exactly
SUPPORTED_SHEETS = [
    "vHost", "vCPU", "vMemory", "vDisk", "vCluster", 
    "vDatastore", "vSC_VMK", "vNIC", "vSnapshot", "vCD"
]

def analyze_sheet(ref_sheet, gen_sheet):
    ref_headers = []
    if ref_sheet.max_row >= 1:
        try:
            ref_headers = [c.value for c in ref_sheet[1] if c.value]
        except IndexError:
            pass # Handle case where max_row says >=1 but sheet is effectively empty or index fails

    gen_headers = []
    if gen_sheet.max_row >= 1:
        try:
             gen_headers = [c.value for c in gen_sheet[1] if c.value]
        except IndexError:
            pass

    ref_set = set(ref_headers)
    gen_set = set(gen_headers)
    
    missing_cols = list(ref_set - gen_set)
    extra_cols = list(gen_set - ref_set)
    
    # Order check
    order_diff = False
    # Only check order for common columns to avoid noise from missing/extra
    common_cols_ref = [h for h in ref_headers if h in gen_set]
    common_cols_gen = [h for h in gen_headers if h in ref_set]
    
    if common_cols_ref != common_cols_gen:
        order_diff = True
        
    ref_rows = max(0, ref_sheet.max_row - 1)
    gen_rows = max(0, gen_sheet.max_row - 1)
    
    return {
        "missing_cols": missing_cols,
        "extra_cols": extra_cols,
        "order_diff": order_diff,
        "ref_rows": ref_rows,
        "gen_rows": gen_rows
    }

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 scripts/compare_rvtools.py <reference_xlsx> <generated_xlsx>")
        sys.exit(1)

    ref_path = Path(sys.argv[1])
    gen_path = Path(sys.argv[2])

    if not ref_path.exists():
        print(f"Error: Reference file not found: {ref_path}")
        sys.exit(1)
        
    if not gen_path.exists():
        print(f"Error: Generated file not found: {gen_path}")
        sys.exit(1)

    try:
        wb_ref = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
        wb_gen = openpyxl.load_workbook(gen_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error opening workbooks: {e}")
        sys.exit(1)

    exit_code = 0
    
    print(f"{'Sheet':<15} | {'Ref Rows':<8} | {'Gen Rows':<8} | {'Missing':<8} | {'Extra':<8} | {'Order Diff':<10}")
    print("-" * 75)

    all_sheets = set(wb_ref.sheetnames) | set(wb_gen.sheetnames)
    
    for sheet_name in sorted(list(all_sheets)):
        if sheet_name not in wb_ref.sheetnames:
            print(f"{sheet_name:<15} | {'N/A':<8} | {'?':<8} | EXTRA SHEET")
            continue
        if sheet_name not in wb_gen.sheetnames:
            print(f"{sheet_name:<15} | {'?':<8} | {'N/A':<8} | MISSING SHEET")
            if sheet_name in SUPPORTED_SHEETS:
                exit_code = 1
            continue

        stats = analyze_sheet(wb_ref[sheet_name], wb_gen[sheet_name])
        
        missing_cnt = len(stats["missing_cols"])
        extra_cnt = len(stats["extra_cols"])
        order_diff = "YES" if stats["order_diff"] else "NO"
        
        print(f"{sheet_name:<15} | {stats['ref_rows']:<8} | {stats['gen_rows']:<8} | {missing_cnt:<8} | {extra_cnt:<8} | {order_diff:<10}")
        
        if sheet_name in SUPPORTED_SHEETS:
            if missing_cnt > 0 or extra_cnt > 0 or stats["order_diff"]:
                print(f"  [FAIL] Issues in supported sheet {sheet_name}:")
                if missing_cnt > 0: print(f"    Missing: {stats['missing_cols']}")
                if extra_cnt > 0: print(f"    Extra: {stats['extra_cols']}")
                if stats["order_diff"]: print(f"    Order mismatch")
                exit_code = 1

    sys.exit(exit_code)

if __name__ == "__main__":
    main()
