#!/usr/bin/env python3
import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


def normalize_header(value: str) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def load_sheet_stats(workbook, sheet_name: str) -> Dict:
    if sheet_name not in workbook.sheetnames:
        return {
            "exists": False,
            "headers": [],
            "row_count": 0,
            "fill_rates": {},
        }

    ws = workbook[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    headers: List[str] = []

    if header_row:
        last_idx = -1
        for idx, val in enumerate(header_row):
            if val is not None and str(val).strip() != "":
                last_idx = idx
        if last_idx >= 0:
            headers = [
                str(val).strip() if val is not None else ""
                for val in header_row[: last_idx + 1]
            ]

    col_count = len(headers)
    non_empty = [0] * col_count
    empty = [0] * col_count
    data_rows = 0

    for row in rows_iter:
        if row is None:
            continue
        if col_count == 0:
            if any(not is_empty(v) for v in row):
                data_rows += 1
            continue

        row_vals = list(row[:col_count])
        if len(row_vals) < col_count:
            row_vals.extend([None] * (col_count - len(row_vals)))

        if not any(not is_empty(v) for v in row_vals):
            continue

        data_rows += 1
        for idx, val in enumerate(row_vals):
            if is_empty(val):
                empty[idx] += 1
            else:
                non_empty[idx] += 1

    fill_rates: Dict[str, Dict[str, float]] = {}
    for idx, header in enumerate(headers):
        total = data_rows
        non = non_empty[idx]
        emp = empty[idx]
        fill_pct = (non / total * 100.0) if total else 0.0
        fill_rates[header] = {
            "non_empty_count": non,
            "empty_count": emp,
            "fill_pct": round(fill_pct, 2),
        }

    return {
        "exists": True,
        "headers": headers,
        "row_count": data_rows,
        "fill_rates": fill_rates,
    }


def build_normalized_matches(
    reference_headers: List[str], generated_headers: List[str]
) -> List[Dict[str, List[str]]]:
    gen_map: Dict[str, List[str]] = {}
    for header in generated_headers:
        norm = normalize_header(header)
        if not norm:
            continue
        gen_map.setdefault(norm, []).append(header)

    matches: List[Dict[str, List[str]]] = []
    gen_headers_set = set(generated_headers)
    for ref_header in reference_headers:
        if ref_header in gen_headers_set:
            continue
        norm = normalize_header(ref_header)
        candidates = gen_map.get(norm, [])
        if candidates:
            matches.append({"reference": ref_header, "generated": candidates})
    return matches


def classify_origin(sheet_name: str, column_name: str) -> str:
    name = column_name.lower()
    if any(token in name for token in ["tools", "guest", "ip address", "ipv4", "ipv6", "dns", "heartbeat"]):
        return "Guest/Tools required (posible pero depende de VMware Tools)"
    if any(token in name for token in ["permission", "privilege", "role", "authorization"]):
        return "Permissions/privileges required (posible pero puede fallar por roles)"

    vm_sheets = {
        "vInfo",
        "vCPU",
        "vMemory",
        "vDisk",
        "vPartition",
        "vNetwork",
        "vCD",
        "vUSB",
        "vSnapshot",
        "vTools",
        "vFileInfo",
        "vRP",
    }
    infra_sheets = {
        "vHost",
        "vCluster",
        "vDatastore",
        "vHBA",
        "vNIC",
        "vSwitch",
        "vPort",
        "dvSwitch",
        "dvPort",
        "vSC_VMK",
        "vMultiPath",
        "vLicense",
        "vSource",
        "vHealth",
    }

    if sheet_name == "vMetaData":
        return "Not available / unclear"
    if sheet_name in vm_sheets:
        return "VM config/runtime/summary (probable)"
    if sheet_name in infra_sheets:
        return "Host/Cluster/Datastore inventory (probable)"
    return "Not available / unclear"


def suggest_location(sheet_name: str, column_name: str) -> str:
    name = column_name.lower()
    if any(token in name for token in ["tools", "guest", "heartbeat"]):
        return "guest.* / runtime.tools"
    if any(token in name for token in ["ip address", "ipv4", "ipv6", "dns"]):
        return "guest.net / config.network"
    if any(token in name for token in ["cpu", "core", "socket"]):
        return "config.hardware / summary"
    if any(token in name for token in ["memory", "ram"]):
        return "config.hardware / summary"
    if any(token in name for token in ["disk", "datastore", "capacity", "provisioned", "used", "free"]):
        return "config.hardware.device / summary.storage / datastore.summary"
    if any(token in name for token in ["snapshot"]):
        return "snapshot / layoutEx"
    if any(token in name for token in ["license"]):
        return "licenseManager"

    sheet_defaults = {
        "vInfo": "summary / config / guest / runtime",
        "vCPU": "config.hardware / summary",
        "vMemory": "config.hardware / summary",
        "vDisk": "config.hardware.device / summary.storage",
        "vPartition": "guest.disk",
        "vNetwork": "guest.net / config.hardware.device",
        "vCD": "config.hardware.device",
        "vUSB": "config.hardware.device",
        "vSnapshot": "snapshot / layoutEx",
        "vTools": "guest / runtime.tools",
        "vFileInfo": "datastoreBrowser",
        "vRP": "resourcePool / summary",
        "vHost": "host.summary / host.config / host.hardware",
        "vCluster": "cluster.summary / configurationEx",
        "vDatastore": "datastore.summary",
        "vHBA": "host.config.storageDevice",
        "vNIC": "host.config.network",
        "vSwitch": "host.config.network",
        "vPort": "host.config.network",
        "dvSwitch": "distributedVirtualSwitch",
        "dvPort": "distributedVirtualSwitch",
        "vSC_VMK": "host.config.network.vnic",
        "vMultiPath": "host.config.storageDevice.multipathInfo",
        "vLicense": "licenseManager",
        "vSource": "serviceInstance.content.about",
        "vHealth": "eventManager",
        "vMetaData": "exporter/runtime",
    }
    return sheet_defaults.get(sheet_name, "unclear")


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_csv(path: Path, rows: List[List[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare RVTools XLSX outputs.")
    parser.add_argument("--reference", required=True, help="Path to RVTools reference XLSX.")
    parser.add_argument(
        "--generated",
        default="out/export.xlsx",
        help="Path to generated XLSX (default: out/export.xlsx).",
    )
    parser.add_argument(
        "--out-dir",
        default="out/compare",
        help="Output directory (default: out/compare).",
    )
    args = parser.parse_args()

    reference_path = Path(args.reference)
    generated_path = Path(args.generated)
    out_dir = Path(args.out_dir)
    ensure_dir(out_dir)

    reference_wb = openpyxl.load_workbook(reference_path, read_only=True, data_only=True)
    generated_wb = openpyxl.load_workbook(generated_path, read_only=True, data_only=True)

    reference_sheets = reference_wb.sheetnames
    generated_sheets = generated_wb.sheetnames
    sheet_order: List[str] = []
    for name in reference_sheets:
        if name not in sheet_order:
            sheet_order.append(name)
    for name in generated_sheets:
        if name not in sheet_order:
            sheet_order.append(name)

    summary: Dict[str, Dict] = {
        "generated": {"path": str(generated_path), "sheet_count": len(generated_sheets)},
        "reference": {"path": str(reference_path), "sheet_count": len(reference_sheets)},
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sheets": {},
    }

    fill_rates_rows = [["sheet", "column", "non_empty_count", "empty_count", "fill_pct"]]
    missing_rows = [["sheet", "column"]]
    extra_rows = [["sheet", "column"]]

    backlog: Dict[str, List[Dict]] = {}
    critical_gaps: List[Tuple[int, str, str, str]] = []
    sheet_gap_counts: List[Tuple[int, str]] = []

    for sheet_name in sheet_order:
        ref_stats = load_sheet_stats(reference_wb, sheet_name)
        gen_stats = load_sheet_stats(generated_wb, sheet_name)

        ref_headers = ref_stats["headers"]
        gen_headers = gen_stats["headers"]

        ref_header_set = set(ref_headers)
        gen_header_set = set(gen_headers)

        missing_columns = [col for col in ref_headers if col not in gen_header_set]
        extra_columns = [col for col in gen_headers if col not in ref_header_set]

        normalized_matches = build_normalized_matches(ref_headers, gen_headers)

        empty_in_generated_not_in_reference: List[str] = []
        empty_in_reference: List[str] = []

        for col in ref_headers:
            ref_fill = ref_stats["fill_rates"].get(col, {})
            ref_non_empty = ref_fill.get("non_empty_count", 0)
            if ref_non_empty == 0:
                empty_in_reference.append(col)

            if col in gen_stats["fill_rates"]:
                gen_non_empty = gen_stats["fill_rates"][col]["non_empty_count"]
                if gen_non_empty == 0 and ref_non_empty > 0:
                    empty_in_generated_not_in_reference.append(col)
                    critical_gaps.append((ref_non_empty, sheet_name, col, "empty"))
            else:
                if ref_non_empty > 0:
                    critical_gaps.append((ref_non_empty, sheet_name, col, "missing"))

        for col, rate in gen_stats["fill_rates"].items():
            fill_rates_rows.append(
                [
                    sheet_name,
                    col,
                    str(rate["non_empty_count"]),
                    str(rate["empty_count"]),
                    f"{rate['fill_pct']:.2f}",
                ]
            )

        for col in missing_columns:
            missing_rows.append([sheet_name, col])
        for col in extra_columns:
            extra_rows.append([sheet_name, col])

        gap_count = len(missing_columns) + len(empty_in_generated_not_in_reference)
        sheet_gap_counts.append((gap_count, sheet_name))

        summary["sheets"][sheet_name] = {
            "reference": {
                "exists": ref_stats["exists"],
                "row_count": ref_stats["row_count"],
                "headers": ref_headers,
            },
            "generated": {
                "exists": gen_stats["exists"],
                "row_count": gen_stats["row_count"],
                "headers": gen_headers,
            },
            "missing_columns": missing_columns,
            "extra_columns": extra_columns,
            "normalized_header_matches": normalized_matches,
            "empty_in_generated_not_in_reference": empty_in_generated_not_in_reference,
            "empty_in_reference": empty_in_reference,
        }

        backlog_items: List[Dict] = []
        for col in missing_columns:
            backlog_items.append(
                {
                    "column": col,
                    "status": "missing",
                    "origin_hint": classify_origin(sheet_name, col),
                    "suggested_location": suggest_location(sheet_name, col),
                    "reference_non_empty_count": ref_stats["fill_rates"].get(col, {}).get(
                        "non_empty_count", 0
                    ),
                    "generated_non_empty_count": 0,
                }
            )
        for col in ref_headers:
            if col not in gen_stats["fill_rates"]:
                continue
            gen_non_empty = gen_stats["fill_rates"][col]["non_empty_count"]
            if gen_non_empty == 0:
                backlog_items.append(
                    {
                        "column": col,
                        "status": "empty_in_generated",
                        "origin_hint": classify_origin(sheet_name, col),
                        "suggested_location": suggest_location(sheet_name, col),
                        "reference_non_empty_count": ref_stats["fill_rates"].get(
                            col, {}
                        ).get("non_empty_count", 0),
                        "generated_non_empty_count": 0,
                    }
                )
        backlog[sheet_name] = backlog_items

    report_lines: List[str] = []
    report_lines.append("# RVTools XLSX comparison report")
    report_lines.append("")
    report_lines.append(f"- reference: {reference_path}")
    report_lines.append(f"- generated: {generated_path}")
    report_lines.append(
        f"- sheets: reference={len(reference_sheets)} generated={len(generated_sheets)}"
    )
    report_lines.append("")

    for sheet_name in sheet_order:
        sheet_data = summary["sheets"][sheet_name]
        report_lines.append(f"## Sheet: {sheet_name}")
        report_lines.append(
            f"- rows: reference={sheet_data['reference']['row_count']} generated={sheet_data['generated']['row_count']}"
        )
        report_lines.append(
            f"- headers: reference={len(sheet_data['reference']['headers'])} generated={len(sheet_data['generated']['headers'])}"
        )
        report_lines.append(
            f"- missing_columns ({len(sheet_data['missing_columns'])}): {', '.join(sheet_data['missing_columns']) or 'none'}"
        )
        report_lines.append(
            f"- extra_columns ({len(sheet_data['extra_columns'])}): {', '.join(sheet_data['extra_columns']) or 'none'}"
        )
        if sheet_data["normalized_header_matches"]:
            report_lines.append("- normalized_header_matches:")
            for match in sheet_data["normalized_header_matches"]:
                report_lines.append(
                    f"  - {match['reference']} -> {', '.join(match['generated'])}"
                )
        else:
            report_lines.append("- normalized_header_matches: none")
        report_lines.append(
            "- empty_in_generated_not_in_reference ({}): {}".format(
                len(sheet_data["empty_in_generated_not_in_reference"]),
                ", ".join(sheet_data["empty_in_generated_not_in_reference"]) or "none",
            )
        )
        report_lines.append(
            "- empty_in_reference ({}): {}".format(
                len(sheet_data["empty_in_reference"]),
                ", ".join(sheet_data["empty_in_reference"]) or "none",
            )
        )
        report_lines.append("")
        report_lines.append("| column | non_empty_count | empty_count | fill_pct |")
        report_lines.append("| --- | --- | --- | --- |")
        gen_fill_rates = load_sheet_stats(generated_wb, sheet_name)["fill_rates"]
        for col in sheet_data["generated"]["headers"]:
            rate = gen_fill_rates.get(col, {"non_empty_count": 0, "empty_count": 0, "fill_pct": 0.0})
            report_lines.append(
                f"| {col} | {rate['non_empty_count']} | {rate['empty_count']} | {rate['fill_pct']:.2f} |"
            )
        report_lines.append("")

    report_path = out_dir / "report.md"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    write_csv(out_dir / "fill_rates_by_sheet.csv", fill_rates_rows)
    write_csv(out_dir / "missing_columns_by_sheet.csv", missing_rows)
    write_csv(out_dir / "extra_columns_by_sheet.csv", extra_rows)

    summary_path = out_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8")

    backlog_path = out_dir / "mapping_backlog.json"
    backlog_payload = {"generated": str(generated_path), "reference": str(reference_path), "sheets": backlog}
    backlog_path.write_text(json.dumps(backlog_payload, indent=2, sort_keys=True), encoding="utf-8")

    critical_gaps.sort(key=lambda x: x[0], reverse=True)
    sheet_gap_counts.sort(key=lambda x: x[0], reverse=True)

    print("Top 10 critical columns (reference filled, missing or empty in generated):")
    for count, sheet, column, reason in critical_gaps[:10]:
        print(f"- {sheet} :: {column} (ref_non_empty={count}, reason={reason})")

    print("")
    print("Top 5 sheets with most gaps:")
    for count, sheet in sheet_gap_counts[:5]:
        print(f"- {sheet} (gap_count={count})")

    print("")
    print(f"Report written to: {report_path}")
    print(f"CSV files written to: {out_dir}")
    print(f"Summary JSON written to: {summary_path}")
    print(f"Mapping backlog written to: {backlog_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
