import openpyxl
import sys

def analyze_export():
    path = "out/export.xlsx"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error opening {path}: {e}")
        return

    print(f"{'Sheet':<15} | {'Rows':<6} | {'Cols':<4} | {'Status':<10}")
    print("-" * 45)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        rows_count = 0
        cols_count = 0
        
        # Check header
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                cols_count = len([c for c in row if c is not None])
            rows_count += 1
            
        data_rows = rows_count - 1
        
        status = "OK"
        if data_rows <= 0:
            status = "VACÍA"
        
        print(f"{sheet_name:<15} | {data_rows:<6} | {cols_count:<4} | {status:<10}")

if __name__ == "__main__":
    analyze_export()
