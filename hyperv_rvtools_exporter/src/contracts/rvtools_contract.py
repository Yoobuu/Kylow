import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook


def _parse_workbook(contract_source: Path) -> Dict[str, Any]:
    """Read the RVTools template and extract sheet/column metadata."""
    wb = load_workbook(contract_source, read_only=True, data_only=True)

    sheets = {}
    sheet_order = []

    for ws in wb.worksheets:
        sheet_order.append(ws.title)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        columns = []
        for col in header_row:
            # Keep exact header value; stringify to avoid None/ints
            columns.append("" if col is None else str(col))

        sheets[ws.title] = {
            "columns": columns,
            "header_row_index": 1
        }

    return {
        "sheet_order": sheet_order,
        "sheets": sheets,
    }


def build_contract(contract_source: Path, contract_json: Path) -> Dict[str, Any]:
    """Build contract from XLSX and persist to JSON."""
    base = _parse_workbook(contract_source)
    contract = {
        **base,
        "source_file": str(contract_source),
        "generated_at": datetime.now(timezone.utc).isoformat()
    }

    contract_json.parent.mkdir(parents=True, exist_ok=True)
    with contract_json.open("w", encoding="utf-8") as f:
        json.dump(contract, f, indent=2)

    return contract


def load_or_build_contract(contract_source: Path, contract_json: Path) -> Dict[str, Any]:
    """
    Load contract if present and newer than the source template,
    otherwise rebuild from the XLSX.
    """
    if contract_json.exists():
        try:
            if contract_json.stat().st_mtime >= contract_source.stat().st_mtime:
                with contract_json.open("r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            # Fall through to rebuild
            pass

    return build_contract(contract_source, contract_json)
