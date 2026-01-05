import argparse
import csv
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import load_workbook
import json


def load_contract(contract_path: Path) -> Dict[str, Any]:
    with contract_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_csv_summary(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"exists": False, "rows": 0, "cols": 0, "headers": []}
    with path.open("r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = sum(1 for _ in reader)
    return {"exists": True, "rows": rows, "cols": len(headers), "headers": headers[:8]}


def read_xlsx_summary(wb, sheet: str) -> Dict[str, Any]:
    if sheet not in wb.sheetnames:
        return {"exists": False, "rows": 0, "cols": 0, "headers": []}
    ws = wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            data_rows += 1
    return {"exists": True, "rows": data_rows, "cols": len(headers), "headers": headers[:8]}


def main():
    parser = argparse.ArgumentParser(description="Debug CSV vs XLSX sheets.")
    parser.add_argument("--out-dir", default="out", help="Output directory containing CSVs and report")
    parser.add_argument("--xlsx", default="out/rvtools_hyperv.xlsx", help="Path to XLSX")
    parser.add_argument("--contract", default=str(Path(__file__).resolve().parent.parent / "contracts" / "rvtools_contract.json"), help="Path to contract JSON")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    xlsx_path = Path(args.xlsx)
    contract_path = Path(args.contract)

    contract = load_contract(contract_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    print(f"Out dir: {out_dir}")
    print(f"XLSX: {xlsx_path}")
    print(f"Contract: {contract_path}")
    print("\nSheet | CSV exists | CSV rows | CSV cols | XLSX rows | XLSX cols | CSV headers (first 8)")

    for sheet in contract.get("sheet_order", []):
        csv_path = out_dir / f"{sheet}.csv"
        csv_sum = read_csv_summary(csv_path)
        xlsx_sum = read_xlsx_summary(wb, sheet)
        print(f"{sheet} | {csv_sum['exists']} | {csv_sum['rows']} | {csv_sum['cols']} | {xlsx_sum['rows']} | {xlsx_sum['cols']} | {csv_sum['headers']}")

    # vInfo detail
    if "vInfo" in wb.sheetnames:
        ws = wb["vInfo"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        sample_rows = []
        for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
            sample_rows.append(row)
        print("\n--- vInfo XLSX detail ---")
        print(f"Headers: {headers[:10]}")
        print(f"Sample rows (first 3 data rows):")
        for r in sample_rows:
            print(r)


if __name__ == "__main__":
    main()
