import argparse
import csv
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook


EMPTY_STRINGS = {"", "null", "none"}


def is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, str):
        if val.strip().lower() in EMPTY_STRINGS:
            return True
        if val.strip() == "":
            return True
    return False


def load_contract(contract_path: Path) -> Dict[str, Any]:
    if not contract_path.exists():
        raise FileNotFoundError(f"Contract not found at {contract_path}")
    with contract_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_csv_sheet(sheet: str, columns: List[str], csv_dir: Path) -> Tuple[int, Dict[str, int], Dict[str, Optional[float]]]:
    path = csv_dir / f"{sheet}.csv"
    counts = {c: 0 for c in columns}
    max_numeric: Dict[str, Optional[float]] = {c: None for c in columns}
    rows = 0
    if not path.exists():
        return rows, counts, max_numeric
    with path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows += 1
            for col in columns:
                val = row.get(col, "")
                if not is_empty(val):
                    counts[col] += 1
                    try:
                        num = float(val)
                        prev = max_numeric[col]
                        if prev is None or num > prev:
                            max_numeric[col] = num
                    except (ValueError, TypeError):
                        pass
    return rows, counts, max_numeric


def read_xlsx_sheet(sheet: str, columns: List[str], wb) -> Tuple[int, Dict[str, int], Dict[str, Optional[float]]]:
    counts = {c: 0 for c in columns}
    max_numeric: Dict[str, Optional[float]] = {c: None for c in columns}
    rows = 0
    if sheet not in wb.sheetnames:
        return rows, counts, max_numeric
    ws = wb[sheet]
    col_len = len(columns)
    for row in ws.iter_rows(min_row=2, max_col=col_len, values_only=True):
        rows += 1
        for idx, col in enumerate(columns):
            val = row[idx] if idx < len(row) else None
            if not is_empty(val):
                counts[col] += 1
                if isinstance(val, (int, float)):
                    prev = max_numeric[col]
                    if prev is None or val > prev:
                        max_numeric[col] = val
    return rows, counts, max_numeric


def process_sheet(sheet: str, columns: List[str], csv_dir: Path, wb) -> Dict[str, Any]:
    rows, counts, max_numeric = read_csv_sheet(sheet, columns, csv_dir)
    source = "csv"
    if rows == 0:
        rows, counts, max_numeric = read_xlsx_sheet(sheet, columns, wb)
        source = "xlsx" if rows > 0 else "none"

    nonempty_list = []
    empty_cols = []
    for col in columns:
        cnt = counts.get(col, 0)
        pct = round((cnt / rows * 100), 2) if rows else 0.0
        if cnt > 0:
            nonempty_list.append({"name": col, "non_empty": cnt, "pct": pct})
        else:
            empty_cols.append(col)

    # Top 10 lists
    top_filled = sorted(nonempty_list, key=lambda x: x["non_empty"], reverse=True)[:10]
    # For emptiest, include empties first then smallest non-empty
    cols_sorted_by_fill = sorted(
        [{"name": c, "non_empty": counts.get(c, 0)} for c in columns],
        key=lambda x: x["non_empty"]
    )[:10]

    warnings = []
    if rows > 0:
        for col, max_val in max_numeric.items():
            if max_val is not None and max_val > 1e9:
                warnings.append(f"{col}: max {max_val} (>1e9) possible units mismatch (bytes vs MiB)")

    return {
        "sheet": sheet,
        "rows": rows,
        "cols_expected": len(columns),
        "non_empty_columns": nonempty_list,
        "empty_columns": empty_cols,
        "top_filled": top_filled,
        "top_empty": cols_sorted_by_fill,
        "source": source,
        "warnings": warnings
    }


def build_summary(sheets_data: List[Dict[str, Any]]) -> Dict[str, Any]:
    total = len(sheets_data)
    with_rows = [s for s in sheets_data if s["rows"] > 0]
    zero_rows = total - len(with_rows)
    avg_pct = 0.0
    if with_rows:
        avg_pct = round(
            sum(len(s["non_empty_columns"]) / s["cols_expected"] * 100 for s in with_rows) / len(with_rows),
            2
        )
    return {
        "total_sheets": total,
        "sheets_with_rows": len(with_rows),
        "sheets_zero_rows": zero_rows,
        "avg_nonempty_cols_pct_on_nonzero": avg_pct
    }


def write_json(report_path: Path, data: Dict[str, Any]):
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def write_markdown(report_path: Path, summary: Dict[str, Any], sheets_data: List[Dict[str, Any]]):
    lines = []
    lines.append(f"# RVTools-style Audit Report")
    lines.append(f"- Generated at: {summary['generated_at']}")
    lines.append(f"- XLSX: {summary['xlsx_path']}")
    lines.append(f"- CSV dir: {summary['csv_dir']}")
    lines.append(
        f"- Sheets: total={summary['summary']['total_sheets']} with_rows={summary['summary']['sheets_with_rows']} "
        f"zero_rows={summary['summary']['sheets_zero_rows']} "
        f"avg_nonempty_cols% (nonzero sheets)={summary['summary']['avg_nonempty_cols_pct_on_nonzero']}"
    )
    lines.append("")

    for sheet_data in sheets_data:
        lines.append(f"## {sheet_data['sheet']}")
        lines.append(
            f"Rows: {sheet_data['rows']} | Cols: {sheet_data['cols_expected']} | "
            f"Non-empty cols: {len(sheet_data['non_empty_columns'])} | Empty cols: {len(sheet_data['empty_columns'])} "
            f"| Source: {sheet_data['source']}"
        )
        if sheet_data["warnings"]:
            lines.append("- Warnings:")
            for w in sheet_data["warnings"]:
                lines.append(f"  - {w}")
        lines.append("- Non-empty columns:")
        if sheet_data["non_empty_columns"]:
            for col in sheet_data["non_empty_columns"]:
                lines.append(f"  - {col['name']}: {col['non_empty']}/{sheet_data['rows']} ({col['pct']}%)")
        else:
            lines.append("  - None")
        lines.append("- Empty columns (0%):")
        if sheet_data["empty_columns"]:
            for col in sheet_data["empty_columns"]:
                lines.append(f"  - {col}")
        else:
            lines.append("  - None")
        lines.append("")

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Audit coverage for RVTools-style Hyper-V export.")
    parser.add_argument("--xlsx", required=True, help="Path to generated rvtools-style XLSX")
    parser.add_argument("--csv-dir", required=True, help="Directory containing per-sheet CSVs")
    parser.add_argument("--contract", default=str(Path(__file__).resolve().parent.parent / "contracts" / "rvtools_contract.json"),
                        help="Path to contract JSON")
    parser.add_argument("--out", default="out/audit_report", help="Output path prefix (without extension)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    csv_dir = Path(args.csv_dir)
    contract_path = Path(args.contract)
    out_prefix = Path(args.out)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found at {xlsx_path}")
    if not csv_dir.exists():
        raise FileNotFoundError(f"CSV directory not found at {csv_dir}")

    contract = load_contract(contract_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    sheets_data = []
    for sheet in contract.get("sheet_order", []):
        cols = contract.get("sheets", {}).get(sheet, {}).get("columns", [])
        sheet_data = process_sheet(sheet, cols, csv_dir, wb)
        sheets_data.append(sheet_data)

    summary = build_summary(sheets_data)
    report_json = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_path": str(xlsx_path),
        "csv_dir": str(csv_dir),
        "contract_path": str(contract_path),
        "summary": summary,
        "sheets": {s["sheet"]: {
            "rows": s["rows"],
            "cols_expected": s["cols_expected"],
            "non_empty_columns": s["non_empty_columns"],
            "empty_columns": s["empty_columns"],
            "top_filled": s["top_filled"],
            "top_empty": s["top_empty"],
            "source": s["source"],
            "warnings": s["warnings"]
        } for s in sheets_data}
    }

    json_path = out_prefix.with_suffix(".json")
    md_path = out_prefix.with_suffix(".md")

    write_json(json_path, report_json)
    summary["generated_at"] = report_json["generated_at"]
    summary["xlsx_path"] = report_json["xlsx_path"]
    summary["csv_dir"] = report_json["csv_dir"]
    summary_full = {"summary": summary, "generated_at": report_json["generated_at"], "xlsx_path": str(xlsx_path), "csv_dir": str(csv_dir)}
    write_markdown(md_path, summary_full, sheets_data)


if __name__ == "__main__":
    main()
